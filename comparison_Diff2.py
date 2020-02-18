""" TEST Program """

import os
# import sys
import glob
import numpy as np
# from tqdm import tqdm
# import datetime
# from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font

path_get = '.\\xls\\*MySQL.xlsx'
path_excel = './xls/'

# 【　str 拡張版　】
def strx(ssP):
	"""
	文字列にする（オブジェクト）文字列を返す Noneは空文字を返す
	"""
	if ssP is None:
		return ''
	elif type(ssP) == str or type(ssP) == int or type(ssP) == np.int64 or type(ssP) == float:
		return str(ssP)
	else:
		return ssP.strftime('%Y/%m/%d %H:%M:%S')

# MySQL側のEXCELリスト獲得
files_sql = glob.glob(path_get)

for fullpath in files_sql:
	mysqlexcelfile = fullpath
	mssqlexcelfile = fullpath[:-10] + 'SQLServer.xlsx'
	diffexcelfile = fullpath[:-10] + 'Diff2.xlsx'

	print('実行中ファイル' + mysqlexcelfile)
	# 開いていたら別名
	try:
		myfile = open(diffexcelfile, "w") # or "a+", whatever you need
		myfile.close()
	except IOError:
		print('警告 {0} が開いているのでファイル名に★追加しまし'.format(diffexcelfile))
		diffexcelfile = fullpath[:-10] + 'Diff2★.xlsx'

	if not os.path.isfile(mssqlexcelfile):
		print('警告：' + fullpath + ' の対象ファイル、' + mssqlexcelfile + ' が存在しないので、スキップします')
		continue



	print('EXCELファイル読み込み中　{0}　'.format(mysqlexcelfile), end='\r', flush=True)
	mywb = openpyxl.load_workbook(mysqlexcelfile)
	print('EXCELファイル読み込み中　{0}　'.format(mssqlexcelfile), end='\r', flush=True)
	mswb = openpyxl.load_workbook(mssqlexcelfile)
	print('     　　　　　　　　　　　　　　　　　　　　　　　　　　　　　　　　　　', end='\r', flush=True)
	myws = mywb['チェックリスト']
	msws = mswb['チェックリスト']
	dfwb = openpyxl.Workbook()
	dfws = dfwb.create_sheet('チェック結果', 0)

	# 差行挿入ループ（２列目がソートされている事）
	myiy = 1
	msiy = 1
	dfiy = 1
	mmix = 2
	P = 0
	mymax = myws.max_row
	msmax = msws.max_row
	dfmax = mymax if mymax > msmax else msmax
	xx = myws.max_column
	a_count = 1

	# 行
	while not (myiy > mymax and msiy > msmax):

		if not (dfiy % 1000) or P == 1 or P == 100:
			print('比較中 : my={0}/{1}  ms={2}/{3}  i={4}　　　　'.format(myiy, mymax, msiy, msmax, dfiy,), end='\r', flush=True)
			if P == 100:
				break
			P = 0
		# ソート用データ取得
		mydata = strx(myws.cell(row=myiy, column=mmix).value)
		msdata = strx(msws.cell(row=msiy, column=mmix).value)
		mydataR = strx(myws.cell(row=myiy, column=mmix + 1).value)
		msdataR = strx(msws.cell(row=msiy, column=mmix + 1).value)

		flg = 0  # 両方セット

		if dfiy == 1: #先頭行はソート対象外
			flg = 0  # 終わり
		elif mydata == '' and msdata == '':
			if myiy > mymax or msiy > msmax:
				P = 100  # 終わり
		elif mydata == '':
			flg = 1  # msだけ
		elif msdata == '':
			flg = 2  # myだけ
		elif myiy == mymax and msiy == msmax:
			flg = 0  # 終わり
		elif myiy == mymax and msiy > msmax:
			flg = 2  # myだけで終わり
		elif msiy == msmax and myiy > mymax:
			flg = 1  # myだけで終わり
		elif mydata > msdata:
			flg = 1  # msだけ
		elif mydata < msdata:
			flg = 2  # myだけ
		elif mydata == msdata and mydataR > msdataR:
			flg = 1  # msだけ
		elif mydata == msdata and  mydataR < msdataR:
			flg = 2  # myだけ

		for x in range(xx):
			if flg == 0:
				mydata = strx(myws.cell(row=myiy, column=x + 1).value)
				msdata = strx(msws.cell(row=msiy, column=x + 1).value)
				dfcell = dfws.cell(row=dfiy, column=x + 1)
				if mydata != msdata:
					dfcell.value = '〓{0}〓{1}'.format(mydata, msdata)
					dfcell.font = Font(color='FF0000')
				else:
					dfcell.value = mydata
			elif flg == 1:
				msdata = strx(msws.cell(row=msiy, column=x + 1).value)
				dfcell = dfws.cell(row=dfiy, column=x + 1)
				dfcell.value = '〓〓{0}'.format(msdata)
				dfcell.font = Font(color='FF8800')
			else: # flg = 2
				mydata = strx(myws.cell(row=myiy, column=x + 1).value)
				dfcell = dfws.cell(row=dfiy, column=x + 1)
				dfcell.value = '〓{0}〓'.format(mydata)
				dfcell.font = Font(color='FF0088')

		if flg == 0:
			myiy += 1
			msiy += 1
			dfiy += 1
		elif flg == 1:
			msiy += 1
			dfiy += 1
			#msmax += 1
			P = 1
		else: # flg = 2
			myiy += 1
			dfiy += 1
			#mymax += 1
			P = 1

	# ファイルセーブ
	print(' 　 　 　 　 　 　 　 　 　 　 　                                       \b\b\b\b\b\b\b\b\b\b\b', end='\r', flush=True)
	print('EXCELファイル書き込み中　{0}　　　　　　'.format(diffexcelfile), end='\r', flush=True)
	dfws.freeze_panes = 'A2'
	dfwb.save(diffexcelfile)

